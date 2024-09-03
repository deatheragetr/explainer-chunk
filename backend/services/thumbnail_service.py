from bson import ObjectId
from typing import Optional, List, Tuple
from PIL import Image, ImageDraw, ImageFont
from PIL.Image import Resampling
import io
import tempfile
import os
import base64
import math
from playwright.async_api import async_playwright
from mypy_boto3_s3.client import S3Client
from config.mongo import TypedAsyncIOMotorDatabase
from config.environment import S3Settings
from db.models.document_uploads import (
    MongoDocumentUpload,
    MongoFileDetails,
    ThumbnailDetails,
)
from pdf2image import convert_from_bytes  # type: ignore[import]
import ebooklib
from ebooklib import epub
from openpyxl import load_workbook
import csv
import json
import markdown
from bs4 import BeautifulSoup
from docx import Document
from utils.file_type_normalizer import supported_file_types
from config.environment import PopplerSettings


poppler_settings = PopplerSettings()


class ThumbnailService:
    def __init__(self, db: TypedAsyncIOMotorDatabase, s3_client: S3Client):
        self.db = db
        self.s3_client = s3_client
        self.s3_settings = S3Settings()
        self.poppler_path = poppler_settings.poppler_path
        self.playwright = None
        self.browser = None
        self.title_font = self.load_font(size=20)
        self.body_font = self.load_font(size=12)

    def load_font(self, size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        # List of font files to try, in order of preference
        font_files = [
            "Arial.ttf",
            "Helvetica.ttf",
            "DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # Common location on Linux
            "/System/Library/Fonts/Helvetica.ttc",  # Common location on macOS
        ]

        for font_file in font_files:
            try:
                return ImageFont.truetype(font_file, size)
            except IOError:
                continue

        # If all else fails, use the default font
        return ImageFont.load_default()

    async def initialize_playwright(self):
        if not self.playwright:
            self.playwright = await async_playwright().start()
            self.browser = await self.playwright.chromium.launch()

    async def close_playwright(self):
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()

    async def generate_and_store_thumbnail(self, document_upload_id: str) -> None:
        document = await self.get_document_upload(document_upload_id)
        if not document:
            raise ValueError(f"Document upload with id {document_upload_id} not found")

        file_details = document["file_details"]
        file_content = await self.get_file_content(file_details)

        normalized_file_type = self.get_normalized_file_type(file_details["file_type"])
        thumbnail = await self.generate_thumbnail(file_content, normalized_file_type)
        thumbnail_key = f"document_upload_thumbnails/{document_upload_id}.png"

        await self.store_thumbnail_in_s3(thumbnail, thumbnail_key)

        thumbnail_url = f"https://{self.s3_settings.s3_document_bucket}.{self.s3_settings.s3_host}/{thumbnail_key}"
        thumbnail_details = ThumbnailDetails(
            file_key=thumbnail_key,
            s3_bucket=self.s3_settings.s3_document_bucket,
            s3_url=thumbnail_url,
        )
        await self.update_document_with_thumbnail(document_upload_id, thumbnail_details)

    async def update_document_with_thumbnail(
        self, document_upload_id: str, thumbnail_details: ThumbnailDetails
    ) -> None:
        await self.db.document_uploads.update_one(
            {"_id": ObjectId(document_upload_id)},
            {"$set": {"thumbnail": thumbnail_details}},
        )

    async def get_document_upload(
        self, document_upload_id: str
    ) -> MongoDocumentUpload | None:
        return await self.db.document_uploads.find_one(
            {"_id": ObjectId(document_upload_id)}
        )

    async def get_file_content(self, file_details: MongoFileDetails) -> bytes:
        response = self.s3_client.get_object(
            Bucket=file_details["s3_bucket"], Key=file_details["file_key"]
        )
        return response["Body"].read()

    def get_normalized_file_type(self, file_type: str) -> str:
        """
        Get the normalized file type based on the supported_file_types mapping.
        If the file type is not found in the mapping, return 'unknown'.
        """
        return next(
            (key for key, value in supported_file_types.items() if value == file_type),
            "unknown",
        )

    async def generate_thumbnail(
        self, file_content: bytes, file_type: str
    ) -> Image.Image:
        if file_type == "pdf":
            return await self.generate_pdf_thumbnail(file_content)
        if file_type == "html":
            html_content = file_content.decode("utf-8")
            thumbnail = await self.generate_html_thumbnail(html_content)
            if thumbnail:
                return thumbnail
            else:
                return self.text_to_image("HTML File (Preview Failed)", file_type)
        elif file_type == "epub":
            return await self.generate_ebook_thumbnail(file_content)
        elif file_type in ["csv", "excel"]:
            return await self.generate_spreadsheet_thumbnail(file_content, file_type)
        elif file_type in ["json", "markdown", "text"]:
            return await self.generate_text_thumbnail(file_content, file_type)
        elif file_type == "word":
            return await self.generate_word_thumbnail(file_content)
        else:
            return await self.generate_default_thumbnail(file_type)

    async def generate_html_thumbnail(self, html_content: str) -> Optional[Image.Image]:
        await self.initialize_playwright()

        try:
            # Create a data URI for the HTML content
            data_uri = f"data:text/html;base64,{base64.b64encode(html_content.encode()).decode()}"

            # Create a new page with strict Content Security Policy
            page = await self.browser.new_page(  # type: ignore
                java_script_enabled=False,  # Disable JavaScript execution
                bypass_csp=False,  # Respect Content Security Policy
            )

            # Set a strict Content Security Policy
            await page.set_extra_http_headers(
                {
                    "Content-Security-Policy": "default-src 'none'; img-src data: http: https:; style-src 'unsafe-inline' http: https:; font-src data: http: https:;"
                }
            )

            # Set viewport size
            await page.set_viewport_size({"width": 1600, "height": 1600})

            # Navigate to the data URI
            await page.goto(data_uri, wait_until="networkidle")

            # Take a screenshot
            screenshot = await page.screenshot(full_page=False)
            await page.close()

            # Process the screenshot to create a 200x200 thumbnail
            with Image.open(io.BytesIO(screenshot)) as img:
                img = img.convert("RGB")

                # Calculate the aspect ratio
                aspect_ratio = img.width / img.height

                if aspect_ratio > 1:  # Width is greater than height
                    new_width = int(200 * aspect_ratio)
                    new_height = 200
                else:  # Height is greater than or equal to width
                    new_width = 200
                    new_height = int(200 / aspect_ratio)

                # Resize the image, maintaining aspect ratio
                img = img.resize((new_width, new_height), Resampling.LANCZOS)

                # Create a new 200x200 white image
                thumbnail = Image.new("RGB", (200, 200), (255, 255, 255))

                # Calculate position to paste resized image centered
                paste_x = (200 - new_width) // 2
                paste_y = (200 - new_height) // 2

                # Paste the resized image onto the white background
                thumbnail.paste(img, (paste_x, paste_y))

                return thumbnail

        except Exception as e:
            print(f"Error generating HTML thumbnail: {e}")
            return None
        finally:
            await self.close_playwright()

    async def generate_pdf_thumbnail(self, file_content: bytes) -> Image.Image:
        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name

        try:
            pages = convert_from_bytes(
                file_content, first_page=1, last_page=1, poppler_path=self.poppler_path
            )
            thumbnail = pages[0].resize((200, 200), Resampling.LANCZOS)
            return thumbnail
        finally:
            os.unlink(temp_file_path)

    async def generate_ebook_thumbnail(self, file_content: bytes) -> Image.Image:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".epub") as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name

        try:
            book = epub.read_epub(temp_file_path)
            for item in book.get_items_of_type(ebooklib.ITEM_COVER):
                if item.media_type.startswith("image/"):
                    cover = Image.open(io.BytesIO(item.content))
                    return cover.resize((200, 200), Resampling.LANCZOS)

            # If no cover image found, generate a default thumbnail
            return await self.generate_default_thumbnail("epub")
        except Exception as e:
            print(f"Error processing epub: {e}")
            return await self.generate_default_thumbnail("epub")
        finally:
            os.unlink(temp_file_path)

    def read_csv(self, file_content: bytes) -> List[List[str]]:
        csv_content = io.StringIO(file_content.decode("utf-8"))
        csv_reader = csv.reader(csv_content)
        return [row for row in csv_reader][:67]  # Limit to first 10 rows

    def read_excel(self, file_content: bytes) -> List[List[str]]:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name

        try:
            wb = load_workbook(filename=temp_file_path, read_only=True)
            sheet = wb.active
            if sheet is None:
                return []  # Return an empty list if there's no active sheet
            return [
                [str(cell.value) if cell.value is not None else "" for cell in row]
                for row in sheet.iter_rows(max_row=67)
            ]
        finally:
            os.unlink(temp_file_path)

    # TODO: REMOVE?
    def create_error_thumbnail(self, error_message: str) -> Image.Image:
        return self.text_to_image(error_message, "Error")

    async def generate_spreadsheet_thumbnail(
        self, file_content: bytes, file_type: str
    ) -> Image.Image:
        if file_type == "csv":
            data = self.read_csv(file_content)
        elif file_type == "excel":
            data = self.read_excel(file_content)
        else:
            return self.create_error_thumbnail("Unsupported spreadsheet type")

        return self.create_spreadsheet_image(data)

    def create_spreadsheet_image(self, data: List[List[str]]) -> Image.Image:
        MAX_WIDTH = 800
        MAX_HEIGHT = 600
        PADDING = 5
        MIN_CELL_WIDTH = 40
        MIN_CELL_HEIGHT = 20

        num_rows = len(data)
        num_cols = len(data[0]) if data else 0

        # Create a temporary image to calculate text dimensions
        temp_img = Image.new("RGB", (1, 1), color="white")
        draw = ImageDraw.Draw(temp_img)

        # Calculate cell dimensions
        col_widths = self.calculate_column_widths(data, draw, MIN_CELL_WIDTH)
        row_height = max(
            MIN_CELL_HEIGHT, self.calculate_row_height(data[0], draw) + PADDING
        )

        # Calculate image dimensions
        img_width = sum(col_widths) + PADDING * (num_cols + 1)
        img_height = row_height * num_rows + PADDING * (num_rows + 1)

        # Scale down if necessary
        scale = min(
            1,
            MIN_CELL_HEIGHT / row_height,
            MAX_WIDTH / img_width,
            MAX_HEIGHT / img_height,
        )

        img_width = math.ceil(img_width * scale)
        img_height = math.ceil(img_height * scale)
        row_height = math.ceil(row_height * scale)
        col_widths = [math.ceil(w * scale) for w in col_widths]

        # Create the actual image
        img = Image.new("RGB", (img_width, img_height), color="white")
        draw = ImageDraw.Draw(img)

        # Draw cells
        y = PADDING
        for row in data:
            x = PADDING
            for i, cell in enumerate(row):
                cell_width = col_widths[i]
                self.draw_cell(draw, x, y, cell_width, row_height, cell, scale)
                x += cell_width + PADDING
            y += row_height + PADDING

        # Resize to 200x200 while maintaining aspect ratio and padding with white
        img = self.resize_and_pad(img, (200, 200))

        return img

    def calculate_column_widths(
        self, data: List[List[str]], draw: ImageDraw.ImageDraw, min_width: int
    ) -> List[int]:
        col_widths = [min_width] * len(data[0])
        for row in data:
            for i, cell in enumerate(row):
                text_width = draw.textbbox((0, 0), str(cell), font=self.body_font)[2]
                col_widths[i] = max(col_widths[i], text_width + 10)  # Add some padding
        return col_widths

    def calculate_row_height(self, row: List[str], draw: ImageDraw.ImageDraw) -> int:
        return max(
            draw.textbbox((0, 0), str(cell), font=self.body_font)[3] for cell in row
        )

    def draw_cell(
        self,
        draw: ImageDraw.ImageDraw,
        x: int,
        y: int,
        width: int,
        height: int,
        content: str,
        scale: float,
    ):
        # Draw cell border
        draw.rectangle([x, y, x + width, y + height], outline="lightgray")

        # Draw cell content
        text = str(content)
        if scale < 1:
            text = self.truncate_text(text, width, draw)
        text_bbox = draw.textbbox((0, 0), text, font=self.body_font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]
        text_x = x + (width - text_width) / 2
        text_y = y + (height - text_height) / 2
        draw.text((text_x, text_y), str(text), fill="black", font=self.body_font)

    def truncate_text(
        self, text: str, max_width: int, draw: ImageDraw.ImageDraw
    ) -> str:
        if draw.textbbox((0, 0), text, font=self.body_font)[2] <= max_width:
            return text
        while (
            len(text) > 1
            and draw.textbbox((0, 0), text + "...", font=self.body_font)[2] > max_width
        ):
            text = text[:-1]
        return text + "..." if len(text) > 1 else text

    def resize_and_pad(self, img: Image.Image, size: Tuple[int, int]) -> Image.Image:
        # Resize image while maintaining aspect ratio
        img.thumbnail(size, Resampling.LANCZOS)

        # Create a white background image
        background = Image.new("RGB", size, (255, 255, 255))

        # Paste the resized image onto the center of the background
        offset = ((size[0] - img.width) // 2, (size[1] - img.height) // 2)
        background.paste(img, offset)

        return background

    async def generate_text_thumbnail(
        self, file_content: bytes, file_type: str
    ) -> Image.Image:
        text_content = file_content.decode("utf-8")
        if file_type == "json":
            try:
                parsed = json.loads(text_content)
                # Pretty print JSON with indentation
                text_content = json.dumps(parsed, indent=2)
                # Limit to approximately 30 lines
                lines = text_content.split("\n")[:30]
                text_content = "\n".join(lines)
            except json.JSONDecodeError:
                text_content = "Invalid JSON content"
        elif file_type == "markdown":
            html = markdown.markdown(text_content)
            text_content = BeautifulSoup(html, "html.parser").get_text()[:500]
        elif file_type == "html":
            text_content = BeautifulSoup(text_content, "html.parser").get_text()[:500]
        elif file_type == "text":
            text_content = text_content[:500]

        return self.text_to_image(text_content, file_type.upper())

    async def generate_word_thumbnail(self, file_content: bytes) -> Image.Image:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name

        try:
            doc = Document(temp_file_path)
            text_content = ""
            for paragraph in doc.paragraphs[:10]:  # Get first 10 paragraphs
                text_content += paragraph.text + "\n"

            text_content = text_content.strip()[:500]  # Limit to 500 characters
            return self.text_to_image(text_content, "Word Document")
        except Exception as e:
            print(f"Error processing Word document: {e}")
            return await self.generate_default_thumbnail("word")
        finally:
            os.unlink(temp_file_path)

    def text_to_image(self, text: str, title: str = "") -> Image.Image:
        img = Image.new("RGB", (200, 200), color="white")
        d = ImageDraw.Draw(img)
        font = ImageFont.load_default()

        # Draw title if provided
        if title:
            d.text((10, 5), title, fill="black", font=font)
            start_y = 25
        else:
            start_y = 10

        # Calculate maximum characters per line
        max_width = 290
        char_width = d.textbbox((0, 0), "A", font=font)[2]
        max_chars = max_width // char_width

        # Draw text
        y = start_y
        for line in text.split("\n"):
            # Preserve indentation
            indent = len(line) - len(line.lstrip())
            indent_space = " " * indent

            # Truncate line if it's too long
            if len(line) > max_chars:
                line = line[: max_chars - 3] + "..."

            d.text((5, y), indent_space + line.lstrip(), fill="black", font=font)
            y += 12

            if y > 195:  # Stop if we've reached the bottom of the image
                break

        return img

    async def generate_default_thumbnail(self, file_type: str) -> Image.Image:
        img = Image.new("RGB", (200, 200), color="lightgray")
        d = ImageDraw.Draw(img)
        font = ImageFont.load_default()
        d.text((10, 10), f"{file_type.upper()} File", fill="black", font=font)
        return img

    async def store_thumbnail_in_s3(
        self, thumbnail: Image.Image, thumbnail_key: str
    ) -> None:
        buffer = io.BytesIO()
        thumbnail.save(buffer, format="PNG")
        buffer.seek(0)

        self.s3_client.put_object(
            Bucket=self.s3_settings.s3_document_bucket,
            Key=thumbnail_key,
            Body=buffer,
            ContentType="image/png",
        )
